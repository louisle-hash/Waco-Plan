import openpyxl, re
from openpyxl.styles import Font, PatternFill, Alignment

SRC = '/Users/louisle/Downloads/WACO-ProductionSchedule-July-2026-Fixed_1.xlsx'
OUT = ('/Users/louisle/Library/CloudStorage/GoogleDrive-production@americanstarus.com/'
       'My Drive/2. PRODUCTION/Mattress/4. Production Plan/WACO/Plan App/'
       'ban-phat-hanh/DANH-MUC-SAN-PHAM-WACO.xlsx')

wb = openpyxl.load_workbook(SRC, data_only=True)

# gom tên sản phẩm ở cột B của mọi sheet Week; nhớ tuần đầu & tuần cuối còn xuất hiện
names = {}
max_week = 0
for sheet in wb.sheetnames:
    if not sheet.lower().startswith('week'):
        continue
    num = int(re.sub(r'\D', '', sheet) or 0)
    max_week = max(max_week, num)
    ws = wb[sheet]
    for r in range(10, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str):
            s = ' '.join(v.split())
            if len(s) > 3:
                lo, hi = names.get(s, (num, num))
                names[s] = (min(lo, num), max(hi, num))

SIZE_ALT = (r'C-?KING|CAL\s*KING|SPLIT\s*KING|TWIN\s*XL|TWIN-XL|TXL|'
            r'FULL\s*XL|FULL-XL|TWIN|FULL|QUEEN|KING')
# cho phép đuôi phụ sau size: "-New", "(VEGA)", ...
TAIL = r'(?P<tail>(?:\s*[-–]\s*\w+|\s*\([^)]*\))*)'
SIZE_RE = re.compile(r'^(?P<base>.+?)\s*[-–]\s*(?P<size>' + SIZE_ALT + r')' + TAIL + r'\s*$', re.I)
SIZE_RE2 = re.compile(r'^(?P<base>.+?)\s+(?P<size>' + SIZE_ALT + r')' + TAIL + r'\s*$', re.I)

SIZE_CANON = {
    'TWIN': 'TWIN', 'FULL': 'FULL', 'QUEEN': 'QUEEN', 'KING': 'KING',
    'TWINXL': 'TWIN-XL', 'TXL': 'TWIN-XL',
    'FULLXL': 'FULL-XL',
    'CKING': 'CAL-KING', 'CALKING': 'CAL-KING',
    'SPLITKING': 'SPLIT-KING',
}

def canon_size(s):
    k = s.upper().strip()
    return SIZE_CANON.get(k.replace(' ', '').replace('-', ''), k)

def slug(text):
    t = text.upper()
    t = t.replace('"', 'IN').replace('.', '')
    t = re.sub(r'\bM\b|\bMATTRESS\b', '', t)      # bỏ chữ thừa lặp ở mọi dòng
    t = re.sub(r'[^A-Z0-9]+', '-', t).strip('-')
    return re.sub(r'-{2,}', '-', t)

rows, skipped = [], []
for full, (first_wk, last_wk) in sorted(names.items()):
    m = SIZE_RE.match(full) or SIZE_RE2.match(full)
    if not m:
        skipped.append(full)
        continue
    base = m.group('base').strip(' -–')
    tail = (m.group('tail') or '').strip(' -–')
    size = canon_size(m.group('size'))
    if len(base) < 3:
        skipped.append(full)
        continue
    rows.append({
        'base': base, 'size': size, 'tail': tail,
        'first': first_wk, 'last': last_wk, 'full': full,
    })

# sinh SKU duy nhất
seen = {}
for r in rows:
    core = slug(r['base'])
    if r['tail']:
        core += '-' + slug(r['tail'])
    sku = f"{core}-{r['size']}"
    if sku in seen:
        seen[sku] += 1
        sku = f"{sku}-{seen[sku]}"
    else:
        seen[sku] = 1
    r['sku'] = sku
    r['category'] = r['base'].split()[0].strip('(),"')

# ghi file Excel đúng cột mà app đang chờ
out = openpyxl.Workbook()
ws = out.active
ws.title = 'San pham'
headers = ['SKU', 'Tên sản phẩm', 'Kích thước', 'Dòng sản phẩm', 'PIC', 'Trạng thái', 'Ghi chú']
ws.append(headers)

hdr_fill = PatternFill('solid', fgColor='041E42')
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color='FFFFFF', name='Arial')
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center')

n_active = 0
for r in sorted(rows, key=lambda x: (x['category'], x['base'], x['size'])):
    active = r['last'] >= max_week
    n_active += active
    note = '' if active else f"Không còn trong kế hoạch từ sau tuần {r['last']} — cân nhắc Lưu trữ"
    if r['tail']:
        note = (note + ' · ' if note else '') + f"biến thể: {r['tail']}"
    ws.append([r['sku'], r['base'], r['size'], r['category'], '', '', note])

ws.freeze_panes = 'A2'
for col, w in zip('ABCDEFG', [34, 40, 12, 16, 10, 14, 52]):
    ws.column_dimensions[col].width = w

out.save(OUT)

print(f'Tổng SKU xuất ra : {len(rows)}')
print(f'  còn dùng (tuần {max_week}) : {n_active}')
print(f'  đã ngừng           : {len(rows) - n_active}')
print(f'Dòng bỏ qua        : {len(skipped)} -> {skipped}')
print(f'Dòng sản phẩm      : {len(set(r["category"] for r in rows))}')
print(f'Đã lưu: {OUT}')
